# portfolio_snapshot_full.py

import yfinance as yf
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

# -------------------------------
# USER SETTINGS
# -------------------------------

portfolio = {
    "AAPL": {"type": "Stock", "units": 25.44, "target_allocation": 0.25, "entry_price": 255.53},
    "BP.L": {"type": "Stock", "units": 10.22, "target_allocation": 0.15, "entry_price": 440.25},
    "JPM": {"type": "Stock", "units": 16.0, "target_allocation": 0.20, "entry_price": 312.47},
    "ULVR.L": {"type": "Stock", "units": 0.84, "target_allocation": 0.15, "entry_price": 4761.50},
    "CSP1.L": {"type": "ETF", "units": 0.06, "target_allocation": 0.15, "entry_price": 55677},
    "VGOV.L": {"type": "Bond", "units": 92.86, "target_allocation": 0.10, "entry_price": 16.153},
}

REBALANCE_THRESHOLD = 2  # % drift threshold for rebalance

# -------------------------------
# FETCH PRICES
# -------------------------------

tickers = list(portfolio.keys())
data = yf.download(tickers, period="14d", interval="1d")["Close"]
latest_prices = data.iloc[-1]
prev_week_prices = data.iloc[-6] if len(data) > 6 else data.iloc[0]

# -------------------------------
# BUILD SNAPSHOT DATAFRAME
# -------------------------------

snapshot_data = []
total_value = sum(latest_prices[t]*portfolio[t]["units"] for t in tickers)

for t in tickers:
    units = portfolio[t]["units"]
    entry = portfolio[t]["entry_price"]
    current_price = latest_prices[t]
    current_value = units * current_price
    initial_value = units * entry
    current_alloc_pct = (current_value / total_value) * 100
    drift = current_alloc_pct - portfolio[t]["target_allocation"]*100
    rebalance = "YES" if abs(drift) > REBALANCE_THRESHOLD else "NO"
    weekly_move_pct = ((current_price - prev_week_prices[t]) / prev_week_prices[t]) * 100
    unrealised_pnl = (current_price - entry) * units

    snapshot_data.append({
        "Asset Name": t,
        "Ticker": t,
        "Asset Type": portfolio[t]["type"],
        "Initial Allocation %": portfolio[t]["target_allocation"]*100,
        "Entry Price (£)": entry,
        "Units": units,
        "Initial Value (£)": round(initial_value,2),
        "New Value (£)": round(current_value,2),
        "Current Allocation %": round(current_alloc_pct,2),
        "Weekly Move (%)": round(weekly_move_pct,2),
        "Unrealised P&L (£)": round(unrealised_pnl,2),
        "Target Allocation (%)": portfolio[t]["target_allocation"]*100,
        "Drift (%)": round(drift,2),
        "Rebalance Needed?": rebalance
    })

snapshot_df = pd.DataFrame(snapshot_data)

# -------------------------------
# SAVE TO EXCEL WITH FORMATTING
# -------------------------------

with pd.ExcelWriter("portfolio_snapshot.xlsx", engine="openpyxl") as writer:
    snapshot_df.to_excel(writer, index=False, sheet_name="Snapshot")
    ws = writer.sheets["Snapshot"]

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Conditional formatting for positive/negative
    for col_letter in ["H", "J", "K", "M"]:  # New Value, Weekly Move, Unrealised P&L, Drift
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(snapshot_df)+1}",
            CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True,
                       fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(snapshot_df)+1}",
            CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True,
                       fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

print("Portfolio snapshot saved to 'portfolio_snapshot.xlsx'")
